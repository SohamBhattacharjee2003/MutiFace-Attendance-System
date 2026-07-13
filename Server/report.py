"""
report.py
Attendance reports a teacher can actually hand in: .xlsx (Excel) or .csv.

Colleges run on spreadsheets. An attendance system that can only show numbers on a screen
gets copied out by hand into Excel anyway, which is where the errors come from — so the
export is not a nice-to-have, it is the last step of the workflow.

Two sheets:
    Summary  — one row per student: attended / held / percent, worst first, and a
               conditional colour so anyone below 75% is visible at a glance.
    Register — the full student x day grid (P / A), i.e. the raw evidence behind
               the summary, because a percentage nobody can audit is worthless.
"""

import io
import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import classes as C

# 75% is the usual college minimum, so it is the line the colours are drawn around.
PASS_MARK = 75.0

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
PRESENT_FILL = PatternFill("solid", fgColor="D5ECDF")
ABSENT_FILL = PatternFill("solid", fgColor="FBD9D4")
LOW_FILL = PatternFill("solid", fgColor="FBD9D4")
OK_FILL = PatternFill("solid", fgColor="D5ECDF")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def workbook(cls, days=90):
    """The full report as an .xlsx in memory."""
    hist = C.history(cls, days)
    wb = Workbook()

    # ── Summary ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = cls["name"]
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A2"] = (f"Classes held: {hist['days_held']}    "
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws["A2"].font = Font(size=9, color="808080")

    headers = ["Roll", "Name", "Attended", "Classes held", "Attendance %", "Status"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center")

    for r, s in enumerate(hist["students"], start=5):
        low = s["enrolled"] and s["percent"] < PASS_MARK
        row = [
            s["roll"], s["name"], s["attended"], s["held"], s["percent"] / 100.0,
            "Not enrolled" if not s["enrolled"] else ("SHORT" if low else "OK"),
        ]
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            if c == 5:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center")
            if c == 6:
                cell.alignment = Alignment(horizontal="center")
                if s["enrolled"]:
                    cell.fill = LOW_FILL if low else OK_FILL
                    cell.font = Font(bold=True, color="C00000" if low else "1E7A45")
    _autosize(ws, [16, 28, 11, 13, 14, 14])
    ws.freeze_panes = "A5"

    # ── Register (the raw grid the summary is derived from) ───────────────────
    ws2 = wb.create_sheet("Register")
    ws2["A1"] = f"{cls['name']} — daily register"
    ws2["A1"].font = Font(bold=True, size=12, color="1F3864")

    head = ["Roll", "Name"] + hist["dates"] + ["%"]
    for c, h in enumerate(head, start=1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", textRotation=90 if c > 2 and c <= len(head) - 1 else 0)

    for r, s in enumerate(hist["students"], start=4):
        ws2.cell(row=r, column=1, value=s["roll"]).border = THIN
        ws2.cell(row=r, column=2, value=s["name"]).border = THIN
        for c, m in enumerate(s["marks"], start=3):
            cell = ws2.cell(row=r, column=c, value="P" if m["present"] else "A")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PRESENT_FILL if m["present"] else ABSENT_FILL
            cell.border = THIN
        pc = ws2.cell(row=r, column=len(head), value=s["percent"] / 100.0)
        pc.number_format = "0.0%"
        pc.alignment = Alignment(horizontal="center")
        pc.font = Font(bold=True)
        pc.border = THIN

    _autosize(ws2, [16, 26] + [4.5] * len(hist["dates"]) + [9])
    ws2.freeze_panes = "C4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def day_csv(cls, day=None):
    """One day's register, as CSV — what a teacher pastes into an email."""
    rep = C.day_report(cls, day)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([cls["name"], rep["date"]])
    w.writerow([])
    w.writerow(["Roll", "Name", "Status", "Time", "Confidence"])
    for r in rep["rows"]:
        w.writerow([r["roll"], r["name"], r["status"],
                    (r["time"] or "")[11:19],
                    f"{r['confidence']:.3f}" if r["confidence"] else ""])
    w.writerow([])
    w.writerow(["Present", rep["present"]])
    w.writerow(["Absent", rep["absent"]])
    w.writerow(["Not enrolled", rep["not_enrolled"]])
    w.writerow(["Attendance %", rep["percent"]])
    return buf.getvalue()


def history_csv(cls, days=90):
    """The whole student x day grid, as CSV."""
    hist = C.history(cls, days)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([cls["name"], f"{hist['days_held']} classes held"])
    w.writerow([])
    w.writerow(["Roll", "Name"] + hist["dates"] + ["Attended", "Held", "Percent"])
    for s in hist["students"]:
        w.writerow([s["roll"], s["name"]]
                   + ["P" if m["present"] else "A" for m in s["marks"]]
                   + [s["attended"], s["held"], s["percent"]])
    return buf.getvalue()
